import os.path
from typing import List
from urllib.parse import urlparse

from openpyxl import load_workbook, Workbook

from publication import *


def write_output(publication: Publication,
                 output_workbook: str,
                 output_sheet: str = 'Sheet1'
                 ):
    """
    Write input Publication's article list to output excel sheet.

    If output workbook does not exist, create it. Else, overwrite
    existing workbook sheet entries.

    Parameters
    ----------
    publication : Publication
        Input Publication to write to output.
    output_workbook : str
        Name of the workbook to be written to.
    output_sheet : str
        Workbook sheet to write to. Defaults to 'Sheet1'.

    Examples
    --------
    index : int
        Article sheet index.
        Written to column 1.
    publication.name : str
        Name of input publication.
        Written to column 2.
    url : str
        Article URL.
        Written to column 3.
    messages : Optional[List[str]]
        String list of debug messages.
        Written to column 4.
    date : str
        Article date.
        Written to column 5.
    num_images : int
        Number of images in article.
        Written to column 6.
    images : Optional[List[Dict[str, str]]]
        List of dictionaries of Caption/Credit pairs.
        Written to columns 7, 8, 9, 10...

    """
    if os.path.isfile(output_workbook):
        print("Opening existing file %s" % output_workbook)
    else:
        print("Creating new file %s" % output_workbook)
        wb = Workbook()
        wb.remove(wb.active)
        wb.create_sheet(output_sheet)
        wb_sheet = wb[output_sheet]
        wb_sheet.cell(row=1, column=1, value='index')
        wb_sheet.cell(row=1, column=2, value='publisher')
        wb_sheet.cell(row=1, column=3, value='article_url')
        wb_sheet.cell(row=1, column=4, value='messages')
        wb_sheet.cell(row=1, column=5, value='date')
        wb_sheet.cell(row=1, column=6, value='num_images')
        wb_sheet.cell(row=1, column=7, value='caption_1')
        wb_sheet.cell(row=1, column=8, value='credit_1')
        wb.save(output_workbook)
        # add labels here
    out_wb = load_workbook(filename=output_workbook)
    out_sheet = out_wb[output_sheet]
    articles = publication.articles
    print("Writing %d articles to %s:%s..." %
          (len(articles), output_workbook, output_sheet))

    for article in articles:
        a = article.as_dictionary
        out_sheet.cell(row=a['index'] + 1, column=1, value=a['index'])
        out_sheet.cell(row=a['index'] + 1, column=2, value=publication.name)
        out_sheet.cell(row=a['index'] + 1, column=3, value=a['url'])
        messages = ""
        for message in a['messages']:
            messages += message + ", "
        messages = messages[:-2]
        out_sheet.cell(row=a['index'] + 1, column=4, value=messages)
        out_sheet.cell(row=a['index'] + 1, column=5, value=a['date'])
        out_sheet.cell(row=a['index'] + 1, column=6, value=a['num_images'])
        for image_index, image in enumerate(a['images']):
            out_sheet.cell(row=a['index'] + 1,
                           column=7 + image_index * 2,
                           value=image['caption']
                           )
            out_sheet.cell(row=a['index'] + 1,
                           column=8 + image_index * 2,
                           value=image['credit']
                           )
    print("Closing and saving %s...\n" % output_workbook)
    out_wb.save(output_workbook)


def print_article(article: Article):
    """
    Print input article's attributes.

    images : Optional[List[Dict[str, str]]]
        List of dictionaries of Caption/Credit pairs.
    date : str
        Article date.
    num_images : int
        Number of images in article.
    url : str
        Article URL.
    index : int
        Article sheet index.
    messages : Optional[List[str]]
        String list of debug messages.

    """
    article = article.as_dictionary
    print("{")
    for k1, v1 in article.items():
        if k1 == 'images':
            print("  [")
            for caption_num, captions in enumerate(v1):
                for k2, v2 in captions.items():
                    print("    " + k2 + " " + str(caption_num + 1) + ":", v2)
            print("  ]")
            continue
        print("  " + k1 + ":", v1)
    print("}")
    print("\n")


def get_workbook_urls(workbook: str,
                      row_start: int,
                      row_end: int,
                      wb_sheet: str = 'Sheet1',
                      col: int = 3
                      ) -> List[str]:
    """
    Return list of URLs extracted from workbook.

    Parameters
    ----------
    workbook : str
        Workbook name.
    row_start : int
        Row to begin parsing.
    row_end : int
        Row to end parsing.
    wb_sheet : str
        Workbook sheet to parse. Defaults to 'Sheet1'.
    col : int
        Column to parse. Defaults to 3

    Returns
    -------
    out : List[str]
        List of extracted urls.

    """
    try:
        wb = load_workbook(filename=workbook)
    except FileNotFoundError:
        print("FileNotFoundError: no such file %s" % workbook)
        return []
    sheet = wb[wb_sheet]
    cells = [sheet.cell(row=row, column=col).value
             for row in range(row_start, row_end)]
    urls = []

    for cell_index, cell in enumerate(cells):
        url_index = cell.find("https://t.co/")
        if url_index == -1:
            urls.append("NO_URL")
            print("No URL found at index %d" % (row_start + cell_index))
            continue

        url = cell[url_index:url_index + 23]
        if not url[13:23].isalnum():
            urls.append("INVALID_URL")
            print("Invalid url found at index %d" % (row_start + cell_index))
        else:
            urls.append(url)

    return urls


def get_domain(url: str) -> str:
    """Return domain of url."""
    return '{uri.netloc}'.format(uri=urlparse(url))
